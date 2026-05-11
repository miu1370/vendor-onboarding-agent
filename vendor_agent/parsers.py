import csv
from pathlib import Path
import openpyxl
import pdfplumber


def _parse_list_field(value):
    if value is None:
        return []
    return [v.strip() for v in str(value).split("\n") if v.strip()]


def parse_intake_xlsx(path):
    wb = openpyxl.load_workbook(path)

    intake = {}
    ws = wb.worksheets[0]
    for row in ws.iter_rows(values_only=True):
        field_key = row[1]
        value = row[3]
        if field_key and field_key not in ("Field Key", "Section"):
            intake[field_key] = value

    # Normalize list fields
    for list_field in ("data_access", "system_integrations", "subprocessors_declared"):
        intake[list_field] = _parse_list_field(intake.get(list_field))

    # Parse document checklist from second sheet
    doc_checklist = {}
    if len(wb.worksheets) > 1:
        ws2 = wb.worksheets[1]
        for row in ws2.iter_rows(values_only=True):
            key, provided = row[0], row[1]
            if key and key not in ("Document Key", "Section", "Inputs are synthetic. Use this sheet to identify provided and missing onboarding materials."):
                doc_checklist[key] = {
                    "provided": bool(provided) if isinstance(provided, bool) else False,
                    "artifact": row[2],
                    "note": row[3],
                }

    intake["document_checklist"] = doc_checklist
    return intake


def parse_quote_csv(path):
    items = []
    with open(path, newline="") as f:
        for row in csv.DictReader(f):
            items.append(dict(row))
    return items


def parse_security_questionnaire(path):
    return Path(path).read_text()


def parse_vendor_email(path):
    return Path(path).read_text()


def parse_contract_pdf(path):
    text = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text += (page.extract_text() or "") + "\n"
    return text.strip()


def load_case(case_id, base_path):
    case_dir = Path(base_path) / "cases" / case_id
    return {
        "case_id": case_id,
        "intake": parse_intake_xlsx(case_dir / f"{case_id}_intake.xlsx"),
        "quote": parse_quote_csv(case_dir / f"{case_id}_quote.csv"),
        "security_questionnaire": parse_security_questionnaire(
            case_dir / f"{case_id}_security_questionnaire.md"
        ),
        "vendor_email": parse_vendor_email(case_dir / f"{case_id}_vendor_email.txt"),
        "contract": parse_contract_pdf(case_dir / f"{case_id}_contract.pdf"),
    }


def load_policies(base_path):
    policies = {}
    for f in sorted((Path(base_path) / "docs").glob("*.md")):
        policies[f.stem] = f.read_text()
    return policies
